"""
    Author:         CaptCorpMURICA
    Project:        100DaysPython
    File:           module3_day32_excel.py
    Creation Date:  6/25/2019, 5:26 PM
    Description:    Python Automation Program 4: Excel Manipulations
"""

import openpyxl

# Even though Excel is not a database tool, multiple sheets or workbooks are often used in this fashion. Additionally,
# many data entry tasks entail copying data from one spreadsheet and pasting it in another. Both of these functions are
# prone to errors. The `openpyxl` package provides enhanced functionality with interfacing with Excel workbooks.
# Load the workbook into a variable. This stores the workbook as a workbook object.
sales_wb = openpyxl.load_workbook('Sales.xlsx')
customers_wb = openpyxl.load_workbook('CustomerDIM.xlsx')
products_wb = openpyxl.load_workbook('ProductDIM.xlsx')

# Load the applicable sheet from each workbook into a variable. This stores the worksheet as a worksheet object.
sales = sales_wb["sales"]
customers = customers_wb["customer_dim"]
products = products_wb["product_dim"]

# The variables act as a named matrix where calling a specific cell with the `.value` method will return a data point.
print(customers["C5"].value)

# Writing to the excel spreadsheet is as easy as calling the cell and setting it equal to the desired value.
# Additionally, the `.value` method can be used for this same functionality.
sales["E1"] = "first_name"
sales["F1"].value = "last_name"
sales["G1"] = "product_name"
sales["H1"] = "product_price"
sales["I1"] = "subtotal"
sales["J1"] = "tax"
sales["K1"] = "total"

# In order to use the customer and product IDs to expand on the data, `for` loops are leveraged to iterate through the
# two sudo-dimension tables to match the identifiers. By enumerating over the contents of the worksheet, each row is
# turned into a tuple. Additionally, starting at `2` ensures the header is skipped in the transformation. The sales tax
# and the final cost can be added once a match is found. This information can easily be obtained through using python
# formulas. Additionally, displaying monies without the proper formatting is not ideal. The `.number_format` method can
# be applied to the `.cell` method. This allows the user to customize the cell formatting. After all of the
# transformations are applied, the workbook is saved to finalize the results.
for i, sr in enumerate(sales.rows, start=2):
    for j, cr in enumerate(customers.rows, start=2):
        if sales[f"B{i}"].value == cr[0].value:
            sales[f"E{i}"] = cr[1].value
            sales[f"F{i}"] = cr[2].value

    for k, pr in enumerate(products.rows, start=2):
        if sales[f"C{i}"].value == pr[0].value:
            sales[f"G{i}"] = pr[1].value
            sales[f"H{i}"] = pr[2].value

    try:
        sales[f"I{i}"] = sales[f"H{i}"].value * sales[f"D{i}"].value
        sales.cell(row=i, column=9).number_format = "$#,##0.00"

        sales[f"J{i}"] = sales[f"I{i}"].value * 0.06
        sales.cell(row=i, column=10).number_format = "$#,##0.00"

        sales[f"K{i}"] = sales[f"I{i}"].value + sales[f"J{i}"].value
        sales.cell(row=i, column=11).number_format = "$#,##0.00"
    except TypeError:
        pass

sales_wb.save("Sales.xlsx")
